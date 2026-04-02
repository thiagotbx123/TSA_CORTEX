"""Generate KPI Pipeline Audit XLSX — auto-dated output to Downloads."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

_DATE = datetime.now().strftime('%Y-%m-%d')
OUTPUT = rf'C:\Users\adm_r\Downloads\AUDIT_KPI_PIPELINE_{_DATE}.xlsx'

# -- Styles ----
CRIT_FILL = PatternFill('solid', fgColor='DC2626')
HIGH_FILL = PatternFill('solid', fgColor='F97316')
MED_FILL  = PatternFill('solid', fgColor='FDE047')
LOW_FILL  = PatternFill('solid', fgColor='BAE6FD')
HDR_FILL  = PatternFill('solid', fgColor='1E293B')
GREEN_FILL = PatternFill('solid', fgColor='DCFCE7')

HDR_FONT    = Font(bold=True, color='FFFFFF', size=11, name='Calibri')
WHITE_FONT  = Font(bold=True, color='FFFFFF', name='Calibri')
BOLD_FONT   = Font(bold=True, name='Calibri')
BASE_FONT   = Font(name='Calibri')
TITLE_FONT  = Font(bold=True, size=16, name='Calibri', color='1E293B')
H2_FONT     = Font(bold=True, size=12, name='Calibri', color='1E293B')
SCORE_FONT  = Font(bold=True, size=22, name='Calibri', color='D97706')

thin = Side(style='thin', color='CBD5E1')
BORDER = Border(top=thin, left=thin, bottom=thin, right=thin)

SEV_FILL = {
    'CRITICAL': CRIT_FILL,
    'HIGH':     HIGH_FILL,
    'MEDIUM':   MED_FILL,
    'LOW':      LOW_FILL,
}

# -- Findings (AUDIT_ENGINE v3.1 full run 2026-04-02) ----
FINDINGS = [
    ('A30-003','CRITICAL','DATA-INTEGRITY',
     'Dual calc_perf with different ETA baselines: merge uses dueDate (current), normalize uses originalEta (first). KPI1 calculation path ambiguous.',
     'merge:316-349 / normalize:112-148',
     'Remove calc_perf from merge. Let normalize be single authority.',
     'HIGH'),
    ('A06-001','HIGH','DEPENDENCIES',
     'Phantom deps: pystray and Pillow imported but not in requirements.txt. Fresh install fails.',
     'kpi_tray.py:44-45',
     'Add pystray>=0.19 and Pillow>=10.0 to requirements.txt',
     'HIGH'),
    ('A08-001','HIGH','TESTING',
     '6 of 8 core modules have zero test coverage. Only merge + normalize pure functions tested.',
     'tests/test_kpi_calculations.py',
     'Add integration tests for orchestrate, build, refresh, tray.',
     'HIGH'),
    ('A04-003','HIGH','SECURITY',
     'HTTP server binds to 0.0.0.0 -- dashboard accessible from any network interface.',
     'kpi_tray.py:194',
     'Change to 127.0.0.1. ngrok handles external access.',
     'HIGH'),
    ('A04-004','HIGH','SECURITY',
     'ngrok exposes dashboard publicly without authentication. Anyone with URL sees KPI data.',
     'kpi_tray.py:75,219',
     'Add --basic-auth user:pass to ngrok command.',
     'HIGH'),
    ('A01-002','HIGH','DATA-INTEGRITY',
     'Customer mapping defined in 3 separate locations. Adding new customer requires 3 edits.',
     'merge:281-296, merge:475-491, normalize:40-67',
     'Consolidate all maps into team_config.py.',
     'HIGH'),
    ('A03-001','HIGH','CODE-QUALITY',
     '2829-line file with embedded HTML/CSS/JS as raw Python string. No syntax highlighting or linting.',
     'build_html_dashboard.py',
     'Extract HTML template to separate .html file. Use placeholder replacement.',
     'HIGH'),
    ('A37-002','HIGH','BUSINESS-LOGIC',
     'Staleness warning uses build date (dateAdd) not API cache freshness (file mtime).',
     'build_html_dashboard.py:63-64',
     'Include _kpi_all_members.json mtime as "Last API refresh" indicator.',
     'HIGH'),
    ('A21-002','HIGH','DOCUMENTATION',
     'File naming inconsistency: build outputs KPI_DASHBOARD.html but upload targets RACCOONS_KPI_DASHBOARD.html.',
     'build_html_dashboard.py:34 / upload:14',
     'Unify to KPI_DASHBOARD.html everywhere.',
     'HIGH'),
    ('A01-003','MEDIUM','DATA-INTEGRITY',
     'State IDs (15 UUIDs) hardcoded without runtime validation against Linear API.',
     'merge_opossum_data.py:24-44',
     'Add startup validation or fallback for unknown states.',
     'HIGH'),
    ('A02-003','MEDIUM','ERROR-HANDLING',
     'upload_dashboard_drive.py has no retry on network failure. Single request, no backoff.',
     'upload_dashboard_drive.py:25-63',
     'Add exponential backoff (max 3 retries).',
     'HIGH'),
    ('A03-002','MEDIUM','CODE-QUALITY',
     'Module-level execution in merge/normalize prevents clean imports. Tests must mock 6+ builtins.',
     'merge_opossum_data.py / normalize_data.py',
     'Wrap main logic in def main(); export pure functions.',
     'HIGH'),
    ('A07-001','MEDIUM','CONFIGURATION',
     'No .env.example file. New developers have no idea what env vars are needed.',
     'Project root',
     'Create .env.example with LINEAR_API_KEY=your_key_here.',
     'HIGH'),
    ('A07-002','MEDIUM','CONFIGURATION',
     'Output path hardcoded to ~/Downloads/ in 4+ files.',
     'build:34, serve:25, tray:72, upload:14',
     'Define OUTPUT_DIR in team_config.py or env var.',
     'MEDIUM'),
    ('A09-002','MEDIUM','BUSINESS-LOGIC',
     'Admin-close detection may misclassify real tickets. Tickets without startedAt forced to On Time.',
     'normalize_data.py:409-420',
     'Add guard: check history event count >= 2.',
     'MEDIUM'),
    ('A11-001','MEDIUM','DATA-INTEGRITY',
     'Partial pipeline failure leaves _dashboard_data.json with unrecalculated perf labels.',
     'orchestrate.py:99-110',
     'Write merge output to staging file; promote after normalize.',
     'MEDIUM'),
    ('A11-002','MEDIUM','DATA-INTEGRITY',
     'Cache freshness 24h warning is print-only. Dashboard can be built from week-old data.',
     'merge_opossum_data.py:228-229',
     'Pass cache age to dashboard; show red banner if >48h.',
     'MEDIUM'),
    ('A13-002','MEDIUM','INFRASTRUCTURE',
     'Tray server log file grows without rotation. No max size, no cleanup.',
     'kpi_tray.py:24',
     'Use RotatingFileHandler (5MB, 3 backups).',
     'MEDIUM'),
    ('A14-001','MEDIUM','INFRASTRUCTURE',
     'No CI/CD pipeline. No automated testing before deployment.',
     'Project root',
     'Add GitHub Actions: pytest on push to master.',
     'MEDIUM'),
    ('A17-001','MEDIUM','ACCESSIBILITY',
     'Dashboard tabs are clickable divs, not button/tab elements. Chart interactions mouse-only.',
     'build_html_dashboard.py JS',
     'Add role=tab, aria-selected; add data table alternative.',
     'MEDIUM'),
    ('A20-001','MEDIUM','MAINTAINABILITY',
     'Bus factor = 1. Single contributor. 59+ data integrity rules with subtle interactions.',
     'Git history',
     'Document top 10 maintenance tasks. Add troubleshooting guide.',
     'HIGH'),
    ('A27-001','MEDIUM','BUSINESS-LOGIC',
     'ETA gaming: set ETA = delivery date guarantees On Time. retroactiveEta flag exists but informational only.',
     'merge:575-577',
     'Show "Organic ETA Accuracy" (excl. retroactive) alongside total.',
     'HIGH'),
    ('A32-001','MEDIUM','BUSINESS-LOGIC',
     'No validation of state transition legality. Unusual paths (Triage->Done) not flagged.',
     'merge_opossum_data.py:96-128',
     'Flag tickets with skipped In Progress step.',
     'MEDIUM'),
    ('A35-002','MEDIUM','BUSINESS-LOGIC',
     'Velocity counts tickets equally regardless of estimate/complexity.',
     'Dashboard JS',
     'Show weighted velocity (sum of estimates) alongside count.',
     'MEDIUM'),
    ('A29-001','MEDIUM','INFRASTRUCTURE',
     'Python 3.14 is bleeding-edge pre-release. Library compat issues possible.',
     'kpi_publish.bat',
     'Test with stable 3.12/3.13. Pin in docs.',
     'MEDIUM'),
    ('A01-004','LOW','CODE-QUALITY',
     'bare except: pass on cache comparison swallows all exceptions including KeyboardInterrupt.',
     'refresh_linear_cache.py:411',
     'Change to except (json.JSONDecodeError, IOError, ValueError).',
     'HIGH'),
    ('A03-003','LOW','CODE-QUALITY',
     '10 variant files in variants/ (5534 lines) appear unused. Dead code.',
     'variants/',
     'Move to archive/ or delete.',
     'HIGH'),
    ('A03-004','LOW','CODE-QUALITY',
     'Magic strings for perf labels (On Time, Late, etc.) used in 4+ files.',
     'merge, normalize, build, tests',
     'Define as constants in team_config.py.',
     'MEDIUM'),
    ('A19-002','LOW','ARCHITECTURE',
     'Two HTTP servers for the same purpose: serve_kpi.py (8787) and kpi_tray.py (8080).',
     'serve_kpi.py / kpi_tray.py',
     'Deprecate serve_kpi.py -- tray is canonical.',
     'LOW'),
]

# -- What's working (I01-I12) ----
WORKING = [
    ('I01', 'Linear API integration with pagination works reliably for standard issue fetch'),
    ('I02', 'History extraction correctly captures ETA changes, state transitions, and assignee changes'),
    ('I03', 'D.LIE scoring engine covers 6 distinct lie dimensions with clear threshold logic'),
    ('I04', 'Gantt chart rendering handles multi-week spans and color-codes by performance tier'),
    ('I05', 'Filter system (person / week / state) correctly narrows all computed tables'),
    ('I06', 'Heatmap generation correctly aggregates weekly on-time rates per person'),
    ('I07', 'orchestrate.py provides single-command full-pipeline execution'),
    ('I08', 'kpi_tray.py system tray app with one-click refresh + serve + ngrok'),
    ('I09', 'Cache-first architecture reduces API calls and allows offline dashboard viewing'),
    ('I10', 'Collapse sections provide clean progressive disclosure of detail rows'),
    ('I11', 'Comprehensive 59-rule D.LIE integrity framework well-documented'),
    ('I12', 'Dual-mode tray (full_refresh vs skip_refresh) enables fast dev iteration'),
]

# -- Top 10 action items ----
ACTIONS = [
    ('1', 'A30-003',          'DATA-INTEGRITY', 'Remove duplicate calc_perf from merge. Single authority in normalize.'),
    ('2', 'A04-003+A04-004',  'SECURITY',       'Bind HTTP to 127.0.0.1; add --basic-auth to ngrok tunnel.'),
    ('3', 'A06-001',          'DEPENDENCIES',    'Add pystray, Pillow, google-auth, etc. to requirements.txt with pinned versions.'),
    ('4', 'A08-001',          'TESTING',         'Add integration tests for orchestrate, build, refresh, upload modules.'),
    ('5', 'A01-002',          'DATA-INTEGRITY',  'Consolidate all customer/person maps into team_config.py.'),
    ('6', 'A03-001+A03-002',  'CODE-QUALITY',    'Extract HTML template from build; wrap merge/normalize in main().'),
    ('7', 'A37-002+A11-002',  'BUSINESS-LOGIC',  'Surface API cache mtime in dashboard; show red banner if >48h stale.'),
    ('8', 'A07-001+A07-002',  'CONFIGURATION',   'Create .env.example; centralize OUTPUT_DIR in config.'),
    ('9', 'A27-001',          'BUSINESS-LOGIC',  'Add "Organic ETA Accuracy" metric excluding retroactive ETAs.'),
    ('10','A17-001+A20-001',  'MAINTAINABILITY', 'Add ARIA roles to dashboard tabs; write troubleshooting guide.'),
]

# -- Category scores ----
CAT_SCORES = [
    ('Security',        40, 100, 'DC2626'),
    ('Data Integrity',  45, 100, 'F97316'),
    ('Error Handling',  55, 100, 'F97316'),
    ('Testing',         30, 100, 'DC2626'),
    ('Dependencies',    50, 100, 'F97316'),
    ('Business Logic',  60, 100, 'FDE047'),
    ('Code Quality',    55, 100, 'F97316'),
    ('Infrastructure',  60, 100, 'FDE047'),
    ('Configuration',   50, 100, 'F97316'),
    ('Accessibility',   45, 100, 'F97316'),
    ('Documentation',   55, 100, 'F97316'),
    ('Maintainability', 45, 100, 'F97316'),
]

# -- Regression Delta (vs 2026-03-26 audit) ----
REGRESSION = [
    ('FIXED', 'C01', 'ngrok stable URL publicly committed'),
    ('FIXED', 'C02', 'Gantt tooltip XSS'),
    ('FIXED', 'H01', 'calc_perf Overdue vs Late divergence'),
    ('FIXED', 'H04', 'merge FileNotFoundError handling'),
    ('FIXED', 'H05', 'Partial fetch silently replaces cache'),
    ('FIXED', 'H07', 'python -m http.server via ngrok'),
    ('FIXED', 'H08', 'STEPS[3] fragile index'),
    ('FIXED', 'H10', '500K-iteration filter rebuild'),
    ('FIXED', 'H11', 'No requirements.txt'),
    ('FIXED', 'H12', 'Activity tab bar breakdown'),
    ('FIXED', 'M01', 'API key inline comments'),
    ('FIXED', 'M05', 'Custom week formula'),
    ('FIXED', 'M09', 'Dedup 40-char collision risk'),
    ('FIXED', 'M15', 'Unknown STATE_NAMES silent'),
    ('FIXED', 'M17', 'netstat port check unreliable'),
    ('NEW', 'A30-003', 'Dual calc_perf ambiguity (escalated to CRITICAL)'),
    ('NEW', 'A37-002', 'Staleness warning uses build date not cache mtime'),
    ('NEW', 'A21-002', 'File naming inconsistency build vs upload'),
    ('NEW', 'A27-001', 'ETA gaming not surfaced'),
    ('NEW', 'A32-001', 'No state transition validation'),
    ('NEW', 'A35-002', 'Velocity unweighted'),
    ('NEW', 'A29-001', 'Python 3.14 pre-release risk'),
    ('NEW', 'A20-001', 'Bus factor = 1'),
    ('NEW', 'A13-002', 'Log rotation missing'),
]


def build_findings_sheet(wb):
    ws = wb.active
    ws.title = 'Findings'
    ws.freeze_panes = 'A2'

    headers = ['ID', 'Severity', 'Category', 'Summary', 'File:Line', 'Recommendation', 'Confidence']
    col_widths = [10, 12, 18, 60, 34, 55, 12]

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER
    ws.row_dimensions[1].height = 20

    for r, f in enumerate(FINDINGS, 2):
        sev = f[1]
        fill = SEV_FILL.get(sev)
        for col, val in enumerate(f, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.border = BORDER
            c.alignment = Alignment(wrap_text=True, vertical='top')
            c.font = BASE_FONT
        sev_cell = ws.cell(row=r, column=2)
        if fill:
            sev_cell.fill = fill
        if sev == 'CRITICAL':
            sev_cell.font = WHITE_FONT
        elif sev in ('HIGH', 'MEDIUM', 'LOW'):
            sev_cell.font = BOLD_FONT
        for col in (1, 3, 4, 5, 6, 7):
            cell = ws.cell(row=r, column=col)
            if r % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='F8FAFC')

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    for r in range(2, len(FINDINGS) + 2):
        ws.row_dimensions[r].height = 44


def build_summary_sheet(wb):
    ws = wb.create_sheet('Summary')

    def cell(row, col, value, font=None, fill=None, align=None):
        c = ws.cell(row=row, column=col, value=value)
        if font:  c.font = font
        if fill:  c.fill = fill
        if align: c.alignment = align
        return c

    ws.merge_cells('A1:G1')
    c = ws.cell(row=1, column=1, value='KPI Pipeline Audit (AUDIT_ENGINE v3.1)')
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.fill = PatternFill('solid', fgColor='F1F5F9')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:G2')
    c2 = ws.cell(row=2, column=1, value=f'Date: {_DATE}   |   Target: Tools/TSA_CORTEX/scripts/kpi/   |   36-Auditor Symbiotic System')
    c2.font = Font(name='Calibri', size=10, color='64748B')
    c2.fill = PatternFill('solid', fgColor='F1F5F9')

    # -- Overall score --
    ws.row_dimensions[4].height = 28
    cell(4, 1, 'OVERALL HEALTH SCORE', font=H2_FONT)
    cell(4, 2, '54 / 100', font=SCORE_FONT,
         align=Alignment(horizontal='center', vertical='center'))
    cell(4, 3, 'YELLOW', font=Font(bold=True, color='D97706', name='Calibri', size=13),
         align=Alignment(horizontal='center', vertical='center'))
    ws.merge_cells('D4:G4')
    cell(4, 4, 'Functional core with critical data integrity gap, security exposure, and low test coverage.',
         font=Font(italic=True, name='Calibri', color='475569'))

    # -- Severity counts --
    cell(6, 1, 'SEVERITY BREAKDOWN', font=H2_FONT)
    hdr_row = 7
    for col, h in enumerate(['Severity', 'Count', 'Share'], 1):
        c = ws.cell(row=hdr_row, column=col, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal='center'); c.border = BORDER

    total = len(FINDINGS)
    crits = sum(1 for f in FINDINGS if f[1] == 'CRITICAL')
    highs = sum(1 for f in FINDINGS if f[1] == 'HIGH')
    meds  = sum(1 for f in FINDINGS if f[1] == 'MEDIUM')
    lows  = sum(1 for f in FINDINGS if f[1] == 'LOW')

    sev_counts = [
        ('CRITICAL', crits, CRIT_FILL, WHITE_FONT),
        ('HIGH',     highs, HIGH_FILL, BOLD_FONT),
        ('MEDIUM',   meds,  MED_FILL,  BOLD_FONT),
        ('LOW',      lows,  LOW_FILL,  BOLD_FONT),
    ]
    for i, (sev, cnt, sfill, sfont) in enumerate(sev_counts, hdr_row + 1):
        c1 = ws.cell(row=i, column=1, value=sev)
        c1.fill = sfill; c1.font = sfont; c1.border = BORDER
        c1.alignment = Alignment(horizontal='center')
        c2 = ws.cell(row=i, column=2, value=cnt)
        c2.font = BOLD_FONT; c2.border = BORDER; c2.alignment = Alignment(horizontal='center')
        c3 = ws.cell(row=i, column=3, value=f'{cnt/total*100:.0f}%')
        c3.border = BORDER; c3.alignment = Alignment(horizontal='center')

    total_row = hdr_row + len(sev_counts) + 1
    ws.cell(row=total_row, column=1, value='TOTAL').font = BOLD_FONT
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal='center')
    ws.cell(row=total_row, column=2, value=total).font = BOLD_FONT
    ws.cell(row=total_row, column=2).alignment = Alignment(horizontal='center')
    for col in range(1, 4):
        ws.cell(row=total_row, column=col).border = BORDER

    # -- Category scores --
    cat_start = total_row + 2
    cell(cat_start, 1, 'CATEGORY SCORES', font=H2_FONT)
    hcat = cat_start + 1
    for col, h in enumerate(['Category', 'Score', 'Out of', 'Grade'], 1):
        c = ws.cell(row=hcat, column=col, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal='center'); c.border = BORDER

    for i, (cat, score, out_of, color) in enumerate(CAT_SCORES, hcat + 1):
        pct = score / out_of
        grade = 'A' if pct >= .9 else 'B' if pct >= .8 else 'C' if pct >= .7 else 'D' if pct >= .6 else 'F'
        cfill = PatternFill('solid', fgColor=color)
        ws.cell(row=i, column=1, value=cat).border = BORDER
        ws.cell(row=i, column=2, value=score).border = BORDER
        ws.cell(row=i, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=i, column=3, value=out_of).border = BORDER
        ws.cell(row=i, column=3).alignment = Alignment(horizontal='center')
        gc = ws.cell(row=i, column=4, value=grade)
        gc.border = BORDER; gc.fill = cfill
        gc.alignment = Alignment(horizontal='center')
        if color in ('DC2626', 'F97316'):
            gc.font = Font(bold=True, color='FFFFFF', name='Calibri')
        else:
            gc.font = BOLD_FONT

    # -- Top 10 action items --
    act_start = hcat + len(CAT_SCORES) + 2
    cell(act_start, 1, 'TOP 10 ACTION ITEMS', font=H2_FONT)
    hact = act_start + 1
    for col, h in enumerate(['#', 'Finding(s)', 'Category', 'Action'], 1):
        c = ws.cell(row=hact, column=col, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal='center'); c.border = BORDER

    for i, (num, refs, cat, action) in enumerate(ACTIONS, hact + 1):
        ws.cell(row=i, column=1, value=num).border = BORDER
        ws.cell(row=i, column=1).alignment = Alignment(horizontal='center')
        rc = ws.cell(row=i, column=2, value=refs)
        rc.border = BORDER
        rc.font = Font(color='4F46E5', bold=True, name='Calibri')
        rc.alignment = Alignment(horizontal='center')
        ws.cell(row=i, column=3, value=cat).border = BORDER
        ac = ws.cell(row=i, column=4, value=action)
        ac.border = BORDER; ac.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[i].height = 30

    # -- What's working well --
    good_start = hact + len(ACTIONS) + 2
    cell(good_start, 1, "WHAT'S WORKING WELL", font=H2_FONT)
    hgood = good_start + 1
    for col, h in enumerate(['ID', 'Observation'], 1):
        c = ws.cell(row=hgood, column=col, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.border = BORDER; c.alignment = Alignment(horizontal='center')

    for i, (iid, obs) in enumerate(WORKING, hgood + 1):
        ic = ws.cell(row=i, column=1, value=iid)
        ic.fill = GREEN_FILL; ic.border = BORDER
        ic.font = Font(bold=True, color='166534', name='Calibri')
        ic.alignment = Alignment(horizontal='center')
        oc = ws.cell(row=i, column=2, value=obs)
        oc.fill = GREEN_FILL; oc.border = BORDER
        oc.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[i].height = 22

    # -- Column widths --
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 65
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12


def build_regression_sheet(wb):
    ws = wb.create_sheet('Regression Delta')
    ws.freeze_panes = 'A2'

    headers = ['Status', 'Finding ID', 'Description']
    col_widths = [10, 14, 70]

    FIXED_FILL = PatternFill('solid', fgColor='DCFCE7')
    NEW_FILL   = PatternFill('solid', fgColor='FEF3C7')

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER

    for r, (status, fid, desc) in enumerate(REGRESSION, 2):
        fill = FIXED_FILL if status == 'FIXED' else NEW_FILL
        sfont = Font(bold=True, color='166534', name='Calibri') if status == 'FIXED' else Font(bold=True, color='92400E', name='Calibri')
        c1 = ws.cell(row=r, column=1, value=status)
        c1.fill = fill; c1.font = sfont; c1.border = BORDER
        c1.alignment = Alignment(horizontal='center')
        c2 = ws.cell(row=r, column=2, value=fid)
        c2.fill = fill; c2.border = BORDER; c2.font = BOLD_FONT
        c2.alignment = Alignment(horizontal='center')
        c3 = ws.cell(row=r, column=3, value=desc)
        c3.fill = fill; c3.border = BORDER
        c3.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[r].height = 22

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.row_dimensions[1].height = 20


# -- Build workbook --
wb = openpyxl.Workbook()
build_findings_sheet(wb)
build_summary_sheet(wb)
build_regression_sheet(wb)

wb.save(OUTPUT)
print(f'Saved: {OUTPUT}')
print(f'Findings: {len(FINDINGS)} rows | Strengths: {len(WORKING)} | Actions: {len(ACTIONS)} | Regression: {len(REGRESSION)}')
